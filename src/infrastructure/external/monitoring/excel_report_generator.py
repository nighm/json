import pandas as pd
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import os
from openpyxl.utils import get_column_letter

def parse_multi_table_csv(csv_file):
    """
    解析多表混合csv，返回每个表的内容（列表的列表）
    返回: [ [表1], [表2], ... ]
    """
    tables = []
    current = []
    with open(csv_file, encoding='utf-8') as f:
        for line in f:
            line = line.rstrip('\n')
            if not line.strip():
                if current:
                    tables.append(current)
                    current = []
            else:
                current.append(line.split(','))
        if current:
            tables.append(current)
    return tables

class ExcelReportGenerator:
    """Excel报告生成器，支持多表csv结构化导入"""
    @staticmethod
    def generate_excel_report(
        csv_file: str,
        trend_img: str,
        dist_img: str,
        dist_explain: str,
        output_file: str,
        summary_info: dict,
        conclusion: str,
        recommendations: list
    ):
        # 1. 解析多表csv
        tables = parse_multi_table_csv(csv_file)
        # 2. 创建Excel工作簿
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = '报告概览'
        row_idx = 1
        # 3. 依次写入每个表
        for table in tables:
            for row in table:
                ws1.append(row)
            ws1.append([])  # 空行分隔
            row_idx += len(table) + 1
        # 4. 结论与建议
        ws1.append(['结论与建议'])
        ws1.append(['结论', conclusion])
        for rec in recommendations:
            ws1.append(['建议', rec])
        # --- Sheet2: 图表 ---
        ws2 = wb.create_sheet('图表')
        # 插入趋势图
        if os.path.exists(trend_img):
            img1 = XLImage(trend_img)
            img1.width = 600
            img1.height = 300
            ws2.add_image(img1, 'A2')
            ws2['A1'] = 'CPU使用率变化趋势图'
        # 插入分布直方图
        if os.path.exists(dist_img):
            img2 = XLImage(dist_img)
            img2.width = 600
            img2.height = 300
            ws2.add_image(img2, 'A22')
            ws2['A21'] = 'CPU使用率分布直方图'
        # 插入中文说明
        if os.path.exists(dist_explain):
            with open(dist_explain, 'r', encoding='utf-8') as f:
                explain = f.read()
            ws2['A43'] = '图表说明：'
            ws2['A44'] = explain
        # 保存Excel
        wb.save(output_file) 